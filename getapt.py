from genericpath import isfile
import requests
import json
import re
import sys
from openpyxl import Workbook
from bs4 import BeautifulSoup
import time as t
import os
import csv
import time
from collections import OrderedDict
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException, TimeoutException
from selenium.webdriver import ActionChains
import zipfile
import chromedriver_autoinstaller

from openpyxl import load_workbook

#미사용 코드1
def download_and_extract_chromedriver(destination_directory):
    
    res = requests.get('https://googlechromelabs.github.io/chrome-for-testing/last-known-good-versions.json').json()
    stable_chromedriver = res['channels']['Stable']['version']
   #url = f"https://edgedl.me.gvt1.com/edgedl/chrome/chrome-for-testing/{stable_chromedriver}/win32/chromedriver-win32.zip"
    url = f"https://storage.googleapis.com/chrome-for-testing-public/{stable_chromedriver}/win64/chromedriver-win64.zip"           
    print("-----------------------------------------------------------------------------------------------------------")
    print(url)
    print("-----------------------------------------------------------------------------------------------------------")
    # 파일 다운로드
    response = requests.get(url)
    zip_filename = os.path.join(destination_directory, "chromedriver-win32.zip")
    
    with open(zip_filename, 'wb') as file:
        file.write(response.content)
    
    # ZIP 파일 압축 해제
    with zipfile.ZipFile(zip_filename, 'r') as zip_ref:
        zip_ref.extractall(destination_directory)
    
    # chromedriver.exe 파일의 경로를 확인
    extracted_chromedriver_path = os.path.join(destination_directory, "chromedriver-win32", "chromedriver.exe")
    
    # 파일이 존재하는지 확인
    if not os.path.exists(extracted_chromedriver_path):
        print(f"{extracted_chromedriver_path} does not exist. Check the zip file contents.")
        return
    
    # chromedriver.exe 파일을 현재 실행 폴더로 복사
    target_chromedriver_path = os.path.join(destination_directory, "chromedriver.exe")
    os.replace(extracted_chromedriver_path, target_chromedriver_path)
    
    # 임시로 다운로드한 ZIP 파일 삭제
    os.remove(zip_filename)
    print("Chromedriver has been successfully extracted and moved to:", target_chromedriver_path)

current_directory = os.getcwd()
#미사용 코드1
#download_and_extract_chromedriver(current_directory)

chromedriver_autoinstaller.install()

wb = load_workbook(filename='upload_by_soonsuboy.xlsx')
ws = wb.active   
city_list=[] #-- 빈 리스트
for cell in ws["A"]: #-- 반복문으로 셀에 D열 값들을 넣는다
    city_list.append(cell.value) #--셀의 값들을 d.col리스트에 추가한다

wb2 = load_workbook(filename='apt_no_upload_by_soonsuboy.xlsx')
ws2 = wb2.active   
excel_apt_no_list=[] #-- 빈 리스트
for cell in ws2["A"]: #-- 반복문으로 셀에 D열 값들을 넣는다
    print ('apt_no' ,cell.value)
    excel_apt_no_list.append(cell.value) #--셀의 값들을 d.col리스트에 추가한다
        

if getattr(sys, 'frozen', False):
    chromedriver_path = os.path.join(sys._MEIPASS, "chromedriver.exe")

import pyautogui
import tkinter as tk
import logging
import datetime

d = datetime.datetime.now()
wrt_dt =d.strftime("%Y.%m.%d")
logger = logging.getLogger()
logger.setLevel(logging.INFO)

formatter = logging.Formatter(u'%(asctime)s [%(levelname)8s] %(message)s')
# StreamHandler
streamingHandler = logging.StreamHandler()
streamingHandler.setFormatter(formatter)
# FileHandler
logfileNm = 'LOG_'+wrt_dt+'.log'
file_handler = logging.FileHandler(logfileNm)
file_handler.setFormatter(formatter)
# RotatingFileHandler
log_max_size = 10 * 1024 * 1024  ## 10MB
log_file_count = 20
logger.addHandler(file_handler)

icon_path = os.path.join(os.path.dirname(__file__), 'soonsuboy_profile.ico')
print (icon_path)


root = tk.Tk()
root.title("Get_apt_by_soonsuboy")
root.geometry("600x700")
if os.path.isfile(icon_path) :
    root.iconbitmap(icon_path)
    
label1 = tk.Label(root,  text="지역 입력하세요 \n ex)강남구 삼성동")
label2 = tk.Label(root,  text="조회할 세대수 입력 \n ex)300 -> 300이상조회")
inputregion = tk.Entry(root, width=15, font=("Courier",8), borderwidth=5)
inputregion.insert(0,'')
inputsaedae = tk.Entry(root, width=15, font=("Courier",8), borderwidth=5)
inputsaedae.insert(0,'200')
label1.grid(column=0, row=0 , padx=5, pady=5 )
label2.grid(column=1, row=0 , padx=5, pady=5 )
inputregion.grid(column=0, row=1 , padx=5, pady=5)
inputsaedae.grid(column=1, row=1 , padx=5, pady=5)

tkpyung=tk.StringVar()
saedae=tk.StringVar()
GrentChk=tk.IntVar()

Rpyung28    = tk.Radiobutton(root , text='23~27평', value = 1  ,variable=tkpyung )
Rpyung2931  = tk.Radiobutton(root , text='28~31평(낀평)', value =2 ,variable=tkpyung )
Rpyung32    = tk.Radiobutton(root , text='32평 이상', value = 3 ,variable=tkpyung )
rentchk    = tk.Checkbutton(root , text='전세만', variable=GrentChk )

Rpyung28.grid(column=3, row=0)
Rpyung2931.grid(column=3, row=1)
Rpyung32.grid(column=3, row=2)
rentchk.grid(column=3, row=3)
Rpyung32.select()

btn7 = tk.Button(root, text='시작', padx=10, pady=5, font=("Courier",10),command=lambda: button_clicked(inputregion,tkpyung,inputsaedae,'regjon'))
btn7.grid(column=1, row=3)

btn8 = tk.Button(root, text='닫기', padx=10, pady=5, font=("Courier",10),command=lambda: button_clolse_clicked())
btn8.grid(column=1, row=4)
## 평대#23~27평#28~31평#31평 이상#범위입력

def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    base_path = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base_path, relative_path)


wall = tk.PhotoImage(file = "soonsuboy_profile.png")
wall_label = tk.Label(image = wall) 
wall_label.grid(column=1, row=5)
labelbottom = tk.Label(root,  text="※평형 조건이 맞지 않는 단지는 최소평수 최저가격 조회됩니다\n※입력한 지역근처 이외의 단지가 나올 수 있습니다")
labelbottom.grid(column=1, row=6 , padx=5, pady=5 )
labelbottom = tk.Label(root,  font=("Courier",8), text="아파트 정보/가격 자동화 for 월부 by 순수보이")
labelbottom.grid(column=1, row=7 , padx=5, pady=5 )
btn9 = tk.Button(root, text='지역엑셀시작', padx=10, pady=5, font=("Courier",10),command=lambda: button_cliked_excel(city_list,tkpyung,inputsaedae))
btn9.grid(column=1, row=8)
btn10 = tk.Button(root, text='아파트번호 엑셀 시작', padx=10, pady=5, font=("Courier",10),command=lambda: button_clicked(inputregion,tkpyung,inputsaedae,'apt_no'))
btn10.grid(column=1, row=9)

city =''

def button_clolse_clicked():
    sys.exit(1)
    

def floorChek (spec):
    for index, value in enumerate(spec):
        realFloor = ''
        totalFloor=''
        print(index , value.text.replace(',','').split())
        parsfloor=value.text.replace(',','').replace('층','').split()
        # print (len(parsfloor[1])) # 3 : 4/6 ,4: 1/23 ,5 :12/44
        if int(len(parsfloor[1])) ==3 :
            realFloor =parsfloor[1][0:1]
            totalFloor=parsfloor[1][2:3]
            print('3 realFloor/totalFloor' ,realFloor,totalFloor)
            logger.info('3 realFloor/totalFloor %s , %s' ,realFloor,totalFloor)
            if realFloor==totalFloor :
                realFloor='top'
                print('top realFloor/totalFloor' ,realFloor,totalFloor)
                logger.info('top realFloor/totalFloor %s , %s' ,realFloor,totalFloor)
        elif int(len(parsfloor[1])) ==4 :
            realFloor =parsfloor[1][0:1]
            totalFloor=parsfloor[1][2:4]
            print('4 realFloor/totalFloor' ,realFloor,totalFloor)
            logger.info('4 realFloor/totalFloor %s , %s' ,realFloor,totalFloor)
            if realFloor==totalFloor :
                realFloor='top'
                print('4 realFloor/totalFloor' ,realFloor,totalFloor)
                logger.info('4 realFloor/totalFloor %s , %s' ,realFloor,totalFloor)
        elif int(len(parsfloor[1])) ==5 :
            realFloor =parsfloor[1][0:2]
            totalFloor=parsfloor[1][3:5]
            print('5 realFloor/totalFloor' ,realFloor,totalFloor)
            logger.info('5 realFloor/totalFloor %s , %s' ,realFloor,totalFloor)
            if realFloor==totalFloor :
                realFloor='top'
                print('5 realFloor/totalFloor' ,realFloor,totalFloor)
                logger.info('5 realFloor/totalFloor %s , %s' ,realFloor,totalFloor)
        else :
            break

    return realFloor 

def pyungChek (spec): # 5: 87/65 , 6:114/84 , 7:120/101
    for index, value in enumerate(spec):
        Gpyung = '0'
        Jpyung = '0'                  
        parspyung=value.text.replace(',','').replace('층','').replace('A','').replace('B','').replace('C','').replace('m²','').split()        
        print('parspyung' , parspyung)
        print (len(parspyung[0]), parspyung[0])
        if int(len(parspyung[0])) ==5 :
            Gpyung =str(int(int(parspyung[0][0:2])/3.3))
            Jpyung =str(int(int(parspyung[0][3:5])/3.3))
            print('5 Gpyung/Jpyung' ,Gpyung,Jpyung)            
            logger.info('5 Gpyung/Jpyung %s , %s' ,Gpyung,Jpyung)      
        elif int(len(parspyung[0])) ==6 :
            Gpyung =str(int(int(parspyung[0][0:3])/3.3))
            Jpyung =str(int(int(parspyung[0][4:6])/3.3))
            print('6 Gpyung/Jpyung' ,Gpyung,Jpyung)            
            logger.info('6 Gpyung/Jpyung %s , %s' ,Gpyung,Jpyung)      
        elif int(len(parspyung[0])) ==7 :
            Gpyung =str(int(int(parspyung[0][0:3])/3.3))
            Jpyung =str(int(int(parspyung[0][5:7])/3.3))
            print('7 Gpyung/Jpyung' ,Gpyung,Jpyung)
            logger.info('7 Gpyung/Jpyung %s , %s' ,Gpyung,Jpyung)      
        elif int(len(parspyung[0])) >7 :
            Gpyung =str(int(int(parspyung[0][0:3])/3.3))            
            print('>7 Gpyung/Jpyung' ,Gpyung,Jpyung)    
            logger.info('>7 Gpyung/Jpyung %s , %s' ,Gpyung,Jpyung)      
        else :
            break          
    
    return Gpyung 

def wayChek (spec): 
    for index, value in enumerate(spec):
        way = ''             
        parsway=value.text.replace(',','').replace('층','').replace('m²','').split()                
        way = parsway[2]        
                
    return way 

def pasprice (p_price): 
    r_price =''
    p_price= p_price.replace(' ','')
    if p_price.find(',') !=-1 :                
        print('pasprice     1' ,p_price , p_price.find(','))
        logger.info('pasprice     1 %s , %s' ,p_price , p_price.find(','))
        r_price= p_price.replace(',','').replace('억','').replace(' ','')
    elif p_price.find(',') == -1 :
        print('pasprice    2' ,p_price ,len(p_price))
        logger.info('pasprice     2 %s , %s' ,p_price ,len(p_price))
        if len(p_price) ==2: #1억            
            r_price= p_price.replace(',','').replace('억','0000').replace(' ','')
            print('pasprice    3' ,r_price)
            logger.info('pasprice     3 %s  ' ,r_price )
        elif len(p_price) ==3: #10억            
            r_price= p_price.replace(',' , '').replace('억','0000').replace(' ','')
            print('pasprice   4' ,r_price)
            logger.info('pasprice     4 %s  ' ,r_price )
        else : #1억500            
            r_price= p_price.replace(',','').replace('억','0').replace(' ','')
            print('pasprice   5' ,r_price)      
            logger.info('pasprice     5 %s  ' ,r_price )
                
    return r_price 





pyungstart=''
pyungend=''
inserted_pyung =''
MJprice0 ='0'
uploadnum=-1
price_list={}
upolad_list =[]



def button_cliked_excel(city_list,tkpyung, tksaedae):    
    for citys in city_list:
        if citys != None :
            logger.info('---------------citys----------------- %s -----------------' ,citys)
            inputregion.delete(0, tk.END)
            inputregion.insert(0,citys)
            button_clicked(inputregion,tkpyung, tksaedae , 'regjon')
    
    

    
        
def button_clicked(region,tkpyung, tksaedae , region_no_gubun):      
    
    apt_no=[]
    
    if tkpyung.get() == '1' :        
            pyungstart = "76"
            pyungend = "89.3"
    elif tkpyung.get() == '2' :        
        pyungstart = "92.6"
        pyungend = "102.5"
    elif tkpyung.get() == '3' :        
        pyungstart = "105.8"
        pyungend = "132.2"
            
            
    if region_no_gubun =='regjon' :
        
        if region.get() =='' :
            pyautogui.alert ("지역을 입력해주세요")
            return
        
        if tksaedae.get() =='' :
            pyautogui.alert ("세대수를 입력해주세요")
            return
            
        print(region.get() , pyungstart , ' ~ ', pyungend)  
        city =region.get()    
        v_saedae    =tksaedae.get()
        # city ="강남구 역삼동"
        # city ="천안시 두정동"

        URL = "https://m.land.naver.com/search/result/" + city
        headers = {"User-Agent" : "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/103.0.0.0 Safari/537.36"}
        res= requests.get(URL ,headers=headers)
        print(URL)
        res.raise_for_status()
        print("응답코드 : ", res.raise_for_status())

        soup = BeautifulSoup(res.text, "lxml")
        # soup.selet("script")

        print(soup.title.get_text())
        print(soup.script.attrs)

        # <script src="https://ssl.pstatic.net/tveta/libs/glad/prod/gfp-core.js"></script>

        # print(soup.find_all('script'))
        print("soup.find_all('script')" , type(soup.find_all('script')))

        data_script = soup.find_all('script')

        #print( data_script[3].text , 'sdw type+2:   ', type(data_script[3].text) )
        # "filter:\" ~"/},"
        # filter 가 포함된 자바스크립트 추출
        data_string =data_script[3].text

        # filter 부분 데이터 추출 
        data_string = data_string.split("filter:")[-1].split("},")[0]
        print( data_string , 'sdw type+1:   ', type(data_string) )

        data_string = data_string.replace('{','').replace('}','').replace(' ','').replace('\n','').replace('"','').replace("'",'').split(',')

        print( data_string , 'sdw type+3:   ', type(data_string) )

        print( 'lat:' ,data_string[0].replace('lat:',''))
        print( 'lon:' ,data_string[1].replace('lon:',''))
        print( 'z:' ,data_string[2].replace('z:',''))
        print( 'cortarNo:' ,data_string[3].replace('cortarNo:',''))
        v1_lat      = data_string[0].replace('lat:','')
        v1_lon      = data_string[1].replace('lon:','')
        v1_z        = data_string[2].replace('z:','')
        v1_cortarNo = data_string[3].replace('cortarNo:','')

        step1Url ='https://m.land.naver.com/cluster/clusterList?cortarNo='+v1_cortarNo+'&rletTpCd=APT&tradTpCd=A1%B1%B2&z='+v1_z+'&lat='+v1_lat+'&lon='+v1_lon+'&pCortarNo=&addon=COMPLEX'
        print( 'step1Url:' ,step1Url)

        cluster1 = requests.get(step1Url ,headers=headers)
        #print(cluster1.text)
        cluster1_parsed = cluster1.text
        cluster1_parsed = cluster1_parsed.split('{"code":"success","data":{"COMPLEX":[')[-1].split(']},')[0]
        cluster1_parsed_lgeo = cluster1_parsed.split('"lgeo":')

        
        apt_saedae_check=[]

        for index2, value2 in enumerate(cluster1_parsed_lgeo) :
            # print("---------------------------------",index2 , value2[1:11] , "sdw type+", type(value2))           
            final_url ='https://m.land.naver.com/cluster/ajax/complexList?itemId='+value2[1:11]+'&mapKey=&lgeo='+value2[1:11]+'&rletTpCd=APT&tradTpCd=A1%B1%B2&z='+v1_z+'&lat='+v1_lat+'&lon='+v1_lon
            # https://m.land.naver.com/cluster/ajax/complexList?itemId='+value2[1:12]+'&mapKey=&lgeo='+value2[1:12]+'&rletTpCd=APT&tradTpCd=A1%B1%B2&z='+v1_z+'&lat='+v1_lat+'&lon=+v1_lon
            print(final_url)
            cluster2 = requests.get(final_url ,headers=headers)
            cluster2_parsed = cluster2.text
            cluster2_parsed = cluster2_parsed.split('{"result":[{')[-1].split('],"more"')[0]
            cluster2_parsed_hscpNo = cluster2_parsed.split('"hscpNo":')

            for index3, value3 in enumerate(cluster2_parsed_hscpNo) :
                apt_no_check ='x'
                apt_saedae_check='x'
                # print(index3 , value3.split('","hscpNm"')[0].replace('"','') , "sdw type+", type(value2))
                apt_no_check = value3.split('","hscpNm"')[0].replace('"','')            
                apt_saedae_check = value3.split('"totHsehCnt":')[-1].split(',"genHsehCnt"')[0]
                if apt_saedae_check !='' and apt_saedae_check != '{"more":false}':
                    
                    if int(apt_saedae_check)>int(v_saedae) :
                        print('2' , apt_saedae_check)
                        if apt_no_check !='' and apt_no_check !='{more:false}':
                            print('3' , apt_saedae_check)
                            apt_no.append(apt_no_check)         
    
    elif region_no_gubun =='apt_no' :
        
        for a_no in excel_apt_no_list:
            if a_no != None :
                print('excel_apt_no_list ' , a_no)
                apt_no.append(a_no)



    driver = webdriver.Chrome()
    URL_APT_NO = "https://new.land.naver.com/complexes/"
    
    # url = 'https://new.land.naver.com/complexes/8621?&a=:JGC:ABYG:APT&b=B1:A1&h=66&i=132'
    

    def get_naver_complex(datacount,id):
        if GrentChk.get() ==0 :
            driver.get(URL_APT_NO + str(id) + "?&a=:JGC:ABYG:APT&b=B1:A1&h="+pyungstart+"&i="+pyungend+"&ad=true")
        elif GrentChk.get() ==1 :
            driver.get(URL_APT_NO + str(id) + "?&a=:JGC:ABYG:APT&b=B1&h="+pyungstart+"&i="+pyungend+"&ad=true")
                
        apt_region=''        
        atp_nm =''
        data = []
        try:
            # if datacount==0 :
            #     data.append('아파트번호')
            #     data.append('부동산 타입')
            #     data.append('시')
            #     data.append('구')
            #     data.append('동')
            #     data.append('아파트명')
            #     data.append('세대수')
            #     data.append('동갯수')
            #     data.append('아파트연식')
            #     data.append('평형정보')
            #     data.append('매물평형')
            #     data.append('매물층')
            #     data.append('전세가')
            #     data.append('전세가')
            #     data.append('매매가')
            #     data.append('매매가')
                
                
            data.append(id)
            category = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CLASS_NAME, "label--category"))).text
            data.append(category)
            
            time.sleep(1)
            index = 1
            for region in WebDriverWait(driver, 10).until(EC.presence_of_all_elements_located((By.CLASS_NAME, "is-selected"))):
                data.append(region.text)
                if index == 3:
                    apt_region = region.text
                    break
                index = index + 1          
          
            title = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.ID, "complexTitle"))).text
            data.append(title)
            atp_nm = apt_region + ' ' +title

            for feature in WebDriverWait(driver, 5).until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, "#summaryInfo > .complex_feature > dd"))):                
                feature =feature.text.replace('㎡','').split('\n')
                data.append(feature[0])                
            
            
            
            
            
            #없는 아파트 검색해서 결과 없을때
            try :
                elem1 = WebDriverWait(driver , 0.5).until(EC.presence_of_element_located ((By.XPATH,"//*[@id='ct']/div[2]/div[2]/div[1]/div/h3" )))
                print('-------------------------------검색 NO DATA----------------------------------------------' )
                # upolad_list.append([title ,'x', 'x','x' ,'x'] )
                logger.info("no result apt : %s",id)                         
            except :
                datacount +=1
                print('-------------------------------',datacount,atp_nm ,'검색 OK----------------------------------------------GrentChk',GrentChk.get() )    
                logger.info('-------------------------------검색 OK %s, %s',datacount,atp_nm)    
        
            try :
                driver.find_element(By.XPATH,"//*[@id='complexOverviewList']/div[2]/div[1]/div[3]/a[3]").click()
                # 정렬한 아파트 가격이 나올때가지 기다린다.
                elem = WebDriverWait(driver , 3).until(EC.presence_of_element_located ((By.XPATH,"//*[@id='articleListArea']" )))        
                # 스크롤링으로 전체 데이터 가져온다.        
                driver.execute_script("document.querySelector('#complexOverviewList > div > div.item_area > div').scrollTop = document.querySelector('#complexOverviewList > div > div.item_area > div').scrollHeight;")
                # driver.execute_script("document.querySelector('.item_list--article').scrollTop = 0;")
                
                time.sleep(0.5)        
                v_datachek = 0 #전세 매매 가격 추출 여부 
                global uploadnum 
                
                if GrentChk.get() == 0 : # 전세트래킹 아닌 경우, 매매전세 최저가 가져오기
                    
                    # 위에가 전센지 매매인지 확인한다. 데이터는 1부터 시작
                    for datanum in range(1,4000) :
                        driver.execute_script("document.querySelector('#complexOverviewList > div > div.item_area > div').scrollTop = document.querySelector('#complexOverviewList > div > div.item_area > div').scrollHeight;")
                        j_Or_m=''
                        MJprice=''
                        spec=[]
                        parsfloor=[]
                        try :
                            #단지내 물건 순서 1번이 낮은가격순
                            dataelem = WebDriverWait(driver , 2).until(EC.presence_of_element_located ((By.XPATH,"//*[@id='articleListArea']/div["+str(datanum)+"]/div[1]/a/div[2]/span[1]" )))
                            #단지내 물건 상세 정보
                            price_list = driver.find_element(By.XPATH,"//*[@id='articleListArea']/div["+str(datanum)+"]/div").text
                            #단지내 물건 전세, 매매 여부확인
                            j_Or_m =driver.find_element(By.XPATH,"//*[@id='articleListArea']/div["+str(datanum)+"]/div[1]/a/div[2]/span[1]").text   
                            MJprice=driver.find_element(By.XPATH,"//*[@id='articleListArea']/div["+str(datanum)+"]/div/a/div[2]/span[2]").text                
                            spec = driver.find_elements(By.XPATH,"//*[@id='articleListArea']/div["+str(datanum)+"]/div[1]/a/div[3]/p[1]/span[1]")
                                                    
                            if j_Or_m =='전세' and v_datachek ==0 :
                                #전세 층수 뽑아내기--------------------------------------------------------                    
                                realFloor = floorChek(spec)
                                Gpyung = pyungChek(spec)
                                way = wayChek(spec)
                                MJprice= pasprice(MJprice)
                                if realFloor != '1' and realFloor != '2' and realFloor != '3' and realFloor != '4' and realFloor != '저'and realFloor != 'top' : 
                                    #평형, 층, 향, 전세가격 추출                        
                                    print(datanum , '기준전세입력--------- atp_nm j_Or_m , realFloor  ,Gpyung ,way,MJprice:',atp_nm , j_Or_m ,realFloor,Gpyung,way,MJprice )
                                    logger.info(' %s , 기준전세입력--------- atp_nm j_Or_m , realFloor  ,Gpyung ,way,MJprice: %s ,%s,%s,%s,%s,%s', datanum ,atp_nm , j_Or_m ,realFloor,Gpyung,way,MJprice )                                
                                    data.append(Gpyung)
                                    data.append(realFloor)
                                    data.append('전세가')
                                    data.append(MJprice)                               
                                    inserted_pyung=Gpyung                               
                                    uploadnum +=1                               
                                    v_datachek =1                               
                                else : # 기준 층 없으면 그냥 최소가격 하나 뽑아내자                        
                                    print(datanum , '최저전세입력--------- atp_nm j_Or_m , realFloor  ,Gpyung ,way,MJprice:',atp_nm , j_Or_m ,realFloor,Gpyung,way,MJprice )
                                    logger.info(' %s , 최저전세입력--------- atp_nm j_Or_m , realFloor  ,Gpyung ,way,MJprice: %s ,%s,%s,%s,%s,%s', datanum ,atp_nm , j_Or_m ,realFloor,Gpyung,way,MJprice )                                
                                    data.append(Gpyung)
                                    data.append(realFloor)
                                    data.append('전세가')
                                    data.append(MJprice)                  
                                    inserted_pyung=Gpyung 
                                    uploadnum +=1
                                    v_datachek =1
                            elif j_Or_m =='매매'  and v_datachek ==1 :
                                #매매 층수 뽑아내기--------------------------------------------------------   
                                realFloor = floorChek(spec)
                                Gpyung = pyungChek(spec)
                                way = wayChek(spec)
                                MJprice= pasprice(MJprice)
                                if inserted_pyung == Gpyung:
                                    print(datanum , '전세입력후 매매입력--------- atp_nm j_Or_m , realFloor  ,Gpyung ,way,MJprice:',atp_nm , j_Or_m ,realFloor,Gpyung,way,MJprice )
                                    logger.info(' %s , 전세입력후 매매입력--------- atp_nm j_Or_m , realFloor  ,Gpyung ,way,MJprice: %s ,%s,%s,%s,%s,%s', datanum ,atp_nm , j_Or_m ,realFloor,Gpyung,way,MJprice )                                
                                    v_datachek =2
                                    data.append('매매가')
                                    data.append(MJprice)
                                    break      
                                                        
                            elif j_Or_m =='매매'  and v_datachek ==0 : #매매밖에 없다면
                                #매매 층수 뽑아내기--------------------------------------------------------   
                                realFloor = floorChek(spec)
                                Gpyung = pyungChek(spec)
                                way = wayChek(spec)
                                MJprice= pasprice(MJprice)                            
                                print(datanum , '매매만--------- atp_nm j_Or_m , realFloor  ,Gpyung ,way,MJprice:',atp_nm , j_Or_m ,realFloor,Gpyung,way,MJprice )
                                logger.info(' %s , 매매만--------- atp_nm j_Or_m , realFloor  ,Gpyung ,way,MJprice: %s ,%s,%s,%s,%s,%s', datanum ,atp_nm , j_Or_m ,realFloor,Gpyung,way,MJprice )
                                data.append(Gpyung)
                                data.append(realFloor)
                                data.append('전세가')
                                data.append(MJprice0)
                                data.append('매매가')
                                data.append(MJprice)
                                uploadnum +=1
                                v_datachek =2    
                                break                    
                            
                        except Exception as e:
                            print("[exception] 발생 : {}".format(e))
                            print('3.0 여기로 오면 에러 또는 완료' )
                            logger.info('3.0 여기로 오면 에러 또는 완료' )
                            break
                
                    #with open('Result_apt_tracking_by_soonsuboy.csv', 'a', newline='', encoding='utf-8') as csvfile:
                    with open('Result_apt_tracking_by_soonsuboy.csv', 'a', newline='') as csvfile:
                        csv_writer = csv.writer(csvfile, delimiter=',',quotechar='\"', quoting=csv.QUOTE_MINIMAL)
                        csv_writer.writerow(data)        
                                            
                elif GrentChk.get() ==1 :
                    print('------------JEONSAE-------------' )
                    # 위에가 전센지 매매인지 확인한다. 데이터는 1부터 시작
                    for datanum in range(1,4000) :
                        driver.execute_script("document.querySelector('#complexOverviewList > div > div.item_area > div').scrollTop = document.querySelector('#complexOverviewList > div > div.item_area > div').scrollHeight;")
                        j_Or_m=''
                        MJprice=''
                        spec=[]
                        parsfloor=[]
                        try :
                            #단지내 물건 순서 1번이 낮은가격순
                            dataelem = WebDriverWait(driver , 2).until(EC.presence_of_element_located ((By.XPATH,"//*[@id='articleListArea']/div["+str(datanum)+"]/div[1]/a/div[2]/span[1]" )))
                            #단지내 물건 상세 정보
                            price_list = driver.find_element(By.XPATH,"//*[@id='articleListArea']/div["+str(datanum)+"]/div").text
                            #단지내 물건 전세, 매매 여부확인
                            j_Or_m =driver.find_element(By.XPATH,"//*[@id='articleListArea']/div["+str(datanum)+"]/div[1]/a/div[2]/span[1]").text   
                            MJprice=driver.find_element(By.XPATH,"//*[@id='articleListArea']/div["+str(datanum)+"]/div/a/div[2]/span[2]").text                
                            spec = driver.find_elements(By.XPATH,"//*[@id='articleListArea']/div["+str(datanum)+"]/div[1]/a/div[3]/p[1]/span[1]")
                            
                            #동을 뽑아내는데 특정 라인 밑으로 내려가면 path가 바뀌므로 없는경우, 새로 넣는다.
                            dong =  driver.find_element(By.XPATH,"//*[@id='articleListArea']/div["+str(datanum)+"]/div[1]/a/div/span").text     

                            if dong =="" :
                                dong =  driver.find_element(By.XPATH,"//*[@id='articleListArea']/div["+str(datanum)+"]/div[1]/a/div[1]/span[2]").text
                                #logger.info(' %s , 동 파악--------- atp_nm j_Or_m , dong  : %s ,%s,%s', datanum ,atp_nm , j_Or_m ,dong)                            
                            
                            if j_Or_m =='전세' :
                                #전세 층수 뽑아내기--------------------------------------------------------                                                    
                                realFloor = floorChek(spec)+'층'
                                Gpyung = pyungChek(spec)
                                way = wayChek(spec)
                                MJprice= pasprice(MJprice)
                               
                                #평형, 층, 향, 전세가격 추출                        
                                print(datanum , 'JEONSAE--------- atp_nm j_Or_m ,dong, realFloor  ,Gpyung ,way,MJprice:',atp_nm , j_Or_m ,dong , realFloor,Gpyung,way,MJprice )
                                logger.info(' %s , JEONSAE--------- atp_nm j_Or_m , dong, realFloor  ,Gpyung ,way,MJprice: %s ,%s,%s,%s,%s,%s,%s', datanum ,atp_nm , j_Or_m ,dong, realFloor,Gpyung,way,MJprice )                                
                                #data.append(Gpyung)
                                #data.append(realFloor)                                                                
                                #data.append(MJprice)         
                                upolad_list.append([dong , realFloor ,Gpyung ,MJprice ] )                      
                        except Exception as e:
                            print("[exception] 발생 : {}".format(e))
                            print('3.0 JEONSAE 여기로 오면 에러 또는 완료' )
                            logger.info('3.0 JEONSAE 여기로 오면 에러 또는 완료' )
                            break  
                
                
                R_upload_list =[]
                R_down_listToStr =[]

                dw= Workbook()
                ds= dw.active
                ds.title = '전세트래킹_'+wrt_dt

                #with open("lent_result_by_soonsuboy.txt", "w", encoding ="utf8") as file:
                        
                for index, value in enumerate(upolad_list):      
                    # print(index , value , ' -----TEST----', value[0],value[1])
                    R_upload_list.append([wrt_dt ,value[0],value[1],value[2],value[3]+"\n"]  )        
                    for ii, vv in enumerate(R_upload_list) :                        
                        if index == ii :                                
                            print(ii , ' '.join(vv) , ' -----TEST1----') # ,' '.join(value) ,type (' '.join(value)))
                            R_down_listToStr.append(' '.join(vv) )
                            #file.writelines(",".join(vv))
                                
                #  위에서 모두 가져온 데이터의 중복을 없애고, 파일에 쓴다.            
                R_down_listToStr = list(OrderedDict.fromkeys(R_down_listToStr))
                print(R_down_listToStr , type(R_down_listToStr))                    
                for ii, vv in enumerate(R_down_listToStr, start=1) :
                    # txt파일로 만들기
                    #file.writelines(vv)
                    # 엑셀로 만들기
                    ds.cell(row=ii ,column=1 , value= vv.split()[0]) #날짜
                    ds.cell(row=ii ,column=2 , value= vv.split()[1]) #아파트
                    ds.cell(row=ii ,column=3 , value= vv.split()[2]) #동
                    ds.cell(row=ii ,column=4 , value= vv.split()[3]) #층
                    ds.cell(row=ii ,column=5 , value= vv.split()[4]) #평
                    ds.cell(row=ii ,column=6 , value= vv.split()[5]) #가격                          

                dw.save("rent_result_by_soonsuboy.xlsx")                
                    
            except :
                print('-------------------------------스크롤 NO DATA----------------------------------------------' )
                # upolad_list.append([atp_nm ,'0', '0','0' ,'0'] )
        
        
        except (NoSuchElementException, TimeoutException) as e:
            print(e)
            return -1

        #with open('Result_apt_tracking_by_soonsuboy.csv', 'a', newline='') as csvfile:
        #    csv_writer = csv.writer(csvfile, delimiter=',',quotechar='\"', quoting=csv.QUOTE_MINIMAL)
        #    csv_writer.writerow(data)

        return 1

                
                
    failed_id = []
    for i, v in enumerate(apt_no) :
        print(i, v )
        if get_naver_complex(i,v) < 0:
            failed_id.append(v)    
        
        # if i == 6:
        #     break;
        
            
    print(failed_id)

    driver.quit()
root.mainloop()
#https://new.land.naver.com/complexes/ + apt_no

# pyinstaller  -w -F --windowed, --noconsole -i="soonsuboy_profile.ico" --add-data="soonsuboy_profile.ico;."  --add-data "soonsuboy_profile.png;." getapt.py
# python -m PyInstaller   -w -F   getapt.py
# pyinstaller  -w -F --windowed, --noconsole getapt.py
# pyinstaller  -w -F  getapt.py



# selenium 은 인터프리터 해결해야함
# 크롬드라이버 오토 인스트롤러 모듈 인식 안되는것 해결하는 pyinstaller
# python -m PyInstaller -w -F --windowed, --noconsole getapt.py